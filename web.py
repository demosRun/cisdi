import re
import os
import json
import datetime
import requests
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, send_file, request

app = Flask(__name__)

@app.route('/')
def home():
    return send_file(os.getcwd() + '/index.html')

@app.route('/edit')
def edit():
    return send_file(os.getcwd() + '/edit.html')

@app.route('/getTableData')
def getTableData():
    outPutData = []
    workbook = load_workbook('file.xlsx')
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        outPutData.append(list(row))
    return json.dumps(outPutData)

@app.route('/saveTable', methods=['POST'])
def saveTable():
    # 获取 JSON 数据
    data = request.json
    print(data)
    # 将数组转换为 DataFrame
    df = pd.DataFrame(data[1:], columns=data[0])

    # 保存为 Excel 文件
    df.to_excel('file.xlsx', index=False)  # index=False 不保存行号
    return json.dumps({"message": "ok"})


def download_file(url, local_filename):
    # 发送 HTTP GET 请求获取文件内容
    with requests.get(url, stream=True) as r:
        r.raise_for_status()  # 检查请求是否成功
        # 打开一个本地文件进行写入
        with open(local_filename, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192): 
                f.write(chunk)  # 将文件内容写入本地文件
    return local_filename


@app.route('/getTable', methods=['GET'])
def getTable():
    # 读取 .xlsx 文件
    df = pd.read_excel('./file.xlsx')

    # 将 DataFrame 转换为 HTML
    # 设置索引从 1 开始
    df.index = range(1, len(df) + 1)
    html_data = df.to_html()
    
    html_data = html_data.replace('NaN', '')
    return html_data

@app.route('/getData', methods=['GET'])
def getData():
    # 加载 Excel 文件
    workbook = load_workbook('file.xlsx')

    # 选择活动工作表
    sheet = workbook.active
    outputData = {
        "节点统计到期图纸": np.zeros(13, dtype=int).tolist(),
        "节点统计到期预算": np.zeros(13, dtype=int).tolist(),
        "节点统计到期采购": np.zeros(13, dtype=int).tolist(),
        "节点统计到期制造": np.zeros(13, dtype=int).tolist(),
        "节点统计到期检验": np.zeros(13, dtype=int).tolist(),
        "节点统计到期发运": np.zeros(13, dtype=int).tolist(),
        "节点统计超期图纸": np.zeros(13, dtype=int).tolist(),
        "节点统计超期预算": np.zeros(13, dtype=int).tolist(),
        "节点统计超期采购": np.zeros(13, dtype=int).tolist(),
        "节点统计超期制造": np.zeros(13, dtype=int).tolist(),
        "节点统计超期检验": np.zeros(13, dtype=int).tolist(),
        "节点统计超期发运": np.zeros(13, dtype=int).tolist(),
        "总已收账款": 0,
        "总应收账款": 0,
        "待收款性质": [0,0,0,0,0,0,0],
        "待收款金额": [0,0,0,0,0,0,0],
        "完成情况": {
            "进行中": 0,
            "发运完成": 0,
            "图纸下达": 0,
            "预算下达": 0,
            "成品检验完成": 0,
            "制造完成": 0,
            "采购合同完成": 0,
        },
        "项目状态": np.zeros(8, dtype=int).tolist()
    }
    def getDateStr (var):
        if (not var or var == '-'):
            # print('有不正确日期!')
            return 12
        if (not isinstance(var, datetime.datetime)):
            var = datetime.datetime.strptime(var, "%Y/%m/%d")
            
        if (var.year == 2024):
            return var.month - 1
        else:
            return 12
    def checkDate(date1, date2):
        if (date2 == '已完成' or date2 == '☑' or date1 == '-' or date2 == '-'):
            return False
        if (not date1):
            return False
        if (not date2):
            return False
        if (not isinstance(date1, datetime.datetime)):
            date1 = datetime.datetime.strptime(date1, "%Y/%m/%d")
        if (not isinstance(date2, datetime.datetime)):
            date2 = datetime.datetime.strptime(date2, "%Y/%m/%d")
        return [date1 < date2]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if (row[0] != None):
            
            try:
                # 项目数量
                outputData["节点统计到期图纸"][getDateStr(row[15])] += 1
                outputData["节点统计到期预算"][getDateStr(row[18])] += 1
                outputData["节点统计到期采购"][getDateStr(row[21])] += 1
                outputData["节点统计到期制造"][getDateStr(row[24])] += 1
                outputData["节点统计到期检验"][getDateStr(row[27])] += 1
                outputData["节点统计到期发运"][getDateStr(row[30])] += 1
                # 延期数量
                if (checkDate(row[15], row[16])):
                    outputData["节点统计超期图纸"][getDateStr(row[15])] += 1
                if (checkDate(row[18], row[19])):
                    outputData["节点统计超期预算"][getDateStr(row[18])] += 1
                if (checkDate(row[21], row[22])):
                    outputData["节点统计超期采购"][getDateStr(row[21])] += 1
                if (checkDate(row[24], row[25])):
                    outputData["节点统计超期制造"][getDateStr(row[24])] += 1
                if (checkDate(row[27], row[28])):
                    outputData["节点统计超期检验"][getDateStr(row[27])] += 1
                if (checkDate(row[30], row[31])):
                    outputData["节点统计超期发运"][getDateStr(row[30])] += 1
                # 收款详情
                if (row[45]):
                    outputData["总已收账款"] += int(row[45])
                if (row[39]):
                    outputData["总应收账款"] += int(row[39])
                if (row[44]):
                    if (row[44] == '预付款'):
                        outputData["待收款性质"][0] += 1
                        outputData["待收款金额"][0] += int(row[39])
                    if (row[44] == '进度款'):
                        outputData["待收款性质"][1] += 1
                        outputData["待收款金额"][1] += int(row[39])
                    if (row[44] == '提货款'):
                        outputData["待收款性质"][2] += 1
                        outputData["待收款金额"][2] += int(row[39])
                    if (row[44] == '到货款'):
                        outputData["待收款性质"][3] += 1
                        outputData["待收款金额"][3] += int(row[39])
                    if (row[44] == '调试款'):
                        outputData["待收款性质"][4] += 1
                        outputData["待收款金额"][4] += int(row[39])
                    if (row[44] == '验收款'):
                        outputData["待收款性质"][5] += 1
                        outputData["待收款金额"][5] += int(row[39])
                    if (row[44] == '质保金'):
                        outputData["待收款性质"][6] += 1
                        outputData["待收款金额"][6] += int(row[39])
                if (row[14]):
                    outputData['项目状态'][0] += 1
                    outputData["完成情况"][row[14]] += 1
                else:
                    outputData['项目状态'][1] += 1
                    outputData["完成情况"]["进行中"] += 1
                # 是否超期
                if (row[38]):
                    outputData['项目状态'][2] += 1
                else:
                    outputData['项目状态'][3] += 1
                # 罚款关注
                if (row[39]):
                    outputData['项目状态'][4] += 1
                if (row[33]):
                    outputData['项目状态'][5] += 1
            except:
                print('数据错误:')
                print(row)
    print(outputData)
    return json.dumps(outputData)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part", 400

    file = request.files['file']
    if file.filename != 'file.xlsx':
        return "No selected file", 400

    # 保存文件
    file.save(f"./{file.filename}")
    return "File uploaded successfully!", 200

if __name__ == '__main__':
    app.run(debug=True, port=5001, host="0.0.0.0")