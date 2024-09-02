import re
import os
import json
import datetime
import requests
import numpy as np
from openpyxl import load_workbook
from flask import Flask, send_file, request

app = Flask(__name__)

@app.route('/')
def home():
    return send_file(os.getcwd() + '/index.html')

@app.route('/table')
def table():
    return send_file(os.getcwd() + '/table.html')

token = 'eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJ7XCJ1c2VySWRcIjpcIjAxNDYzMkBjaXNkaS5jb20uY25cIn0iLCJzdWIiOiLnmbvlvZV0b2tlbiIsImlzcyI6IumXqOaItyIsImlhdCI6MTcyMTM1NDgxMywiZXhwIjoxNzIxMzgzOTEzfQ.ebY4gZY72s-gtPZabJL-yGIfJoQGUxexw-2wQ5iWxH8'
cookie = 'OAMAuthnCookie_ccis-oamgate.cisdi.com.cn:7777=1c82d8a5105dd3c467f0b2819fccf967012911b0%7EVcFBPnASueboBby1RBx5wrxChUkyEEEpyi1RQPefur6GonW1lwF7EEBjk%2FA8eApb8K1sfPwVow%2FIyHxApxPLkGRhONLoncJ6EXQHdpomeMrTkT%2Bjz4e%2BRJ9veOhetJ4q5U35wwI6vYsZP3J5XUAqKT5pxRpxwQ6YOeaC%2FrxcpnS28qSk8b0fA5aE%2BgyWTDXAF1Hz1pGN5sIP1hpAd%2F4WCyhNJn%2Bx%2BzxH%2BHBeUPQ3Aeo%2FZ%2F8CXd8gx12g34A8lNuL8N4JcQvGAkH%2BV%2BU4Q5Z6T%2Fxry6O4m%2BdvGKkvuk9K3OLVaEjyXh7GuQwufoVov%2BN%2BRCTa62QvmG3J8GkzQcVSbdJ3GdG6rgMm3PGKjNvGWYU0v5q4Yf1gscDAFV4OlsqgzMuVxRzFXsi0rAqZ3SPwh0FR3xdcPhGXnhynEsy82%2FeWNrr8OzqABfCVaRtR%2B6z57i%2FlMQj5Ng0eXzz7LSz4qRV0Uim8mZuYqm8pqYCxIRZDU%2F15Q1HSOtG1QHtn4ybtb7u16MC3WMxvn6q5EZ5%2B75WhQM4IQFlFLGFT0ivHNGF2MNEkulzDs7grKLXk673WzlidebA3aECt2MSdZCkoP3YuVNRTL6cyywZX1pP2v0E6cUrJ%2FKjnvdcntdSBWsDZ9g54ERmRfAMiXX96gtXcDfTGhzzg5HGv%2BaFx9Sj5QkHSWSYsDeHShy8TkmcTeGYBeAbeysk3uGZNkJWjGpyTthCMsyETXRk7tdDacDeADDUNrFAK5EOMYK7BjUXu9du%2FURdlSwOecb0AzYyKAGgccA%3D%3D; token=eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJ7XCJ1c2VySWRcIjpcIjAxNDYzMkBjaXNkaS5jb20uY25cIn0iLCJzdWIiOiLnmbvlvZV0b2tlbiIsImlzcyI6IumXqOaItyIsImlhdCI6MTcyMTM1NDgxMywiZXhwIjoxNzIxMzgzOTEzfQ.ebY4gZY72s-gtPZabJL-yGIfJoQGUxexw-2wQ5iWxH8; ccis_sso_token=eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiJ7XCJ1c2VySWRcIjpcIjAxNDYzMkBjaXNkaS5jb20uY25cIn0iLCJzdWIiOiJjY2lz6aqM6K-BdG9rZW4iLCJpc3MiOiLpl6jmiLciLCJpYXQiOjE3MjEzNTQ4MTMsImV4cCI6MTcyMTM4MzkxM30.0hiJMbd32AG1aWvkG2FzEAzMigymBnRKzO_kwjqp_eg; oid=82BD2971CBFC45B3BF10B6E2819EF85F; formatedUserName=014632@cisdi.com.cn; last_login_user=014632; sto-id-24862=GAGEFIAKGBBO; workCode=014632; userId=014632@cisdi.com.cn; portal_session=NjY2; PRD=tnM0nABM9OHtNL0oxeXWPIet:S'

def loadExcel():
    # 加载 Excel 文件
    workbook = load_workbook('file.xlsx')
    sheet = workbook.active
    项目号 = sheet.cell(row=3, column=4).value
    项目名称 = sheet.cell(row=2, column=4).value
    业主名称 = sheet.cell(row=2, column=11).value
    print(项目号 + ':' + 项目名称)

    # 获取从第四行开始的所有内容
    data = []
    # 定义替换函数
    def remove_zero_time(time):
        if (isinstance(time, datetime.datetime)):
            return time.strftime('%Y/%m/%d')
        return time
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if (row[0]):
            lestTemp = [项目名称, row[3], 项目号, "", re.findall(r'\d+', 项目名称)[0], 业主名称, row[7],row[6],row[8], None, None, None, row[12], None,None,None,row[15],None,None,row[16],None,None, row[17],None,None,row[18],None,None,row[19],None,None, row[20]]
            lestTemp = list(map(remove_zero_time, lestTemp))
            data.append(lestTemp)
    
    
    # 加载 Excel 文件
    workbook = load_workbook('main.xlsx')

    # 选择活动工作表
    sheet = workbook.active

    for new_row in data:
        # 在第三行添加一行数据，注意openpyxl的行和列索引从1开始
        sheet.insert_rows(3)
        for col_num, value in enumerate(new_row, start=1):
            sheet.cell(row=3, column=col_num, value=value)

    # 保存文件
    workbook.save('main.xlsx')

def download_file(url, local_filename):
    # 发送 HTTP GET 请求获取文件内容
    with requests.get(url, stream=True) as r:
        r.raise_for_status()  # 检查请求是否成功
        # 打开一个本地文件进行写入
        with open(local_filename, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192): 
                f.write(chunk)  # 将文件内容写入本地文件
    return local_filename

def downFile(fileid):
    url = 'http://oa.cisdi.com.cn:8080/weaver/weaver.file.FileDownload?f_weaver_belongto_userid=18694&f_weaver_belongto_usertype=0&fileid=' + fileid + '&download=1&requestid=3415521&desrequestid=0&fromrequest=1'
    local_filename = 'file.xlsx'
    download_file(url, local_filename)
    print(f"Downloaded {local_filename}")


def getPage(url):
    headers = {
    'token': token,
    'Cookie': cookie
    }

    response = requests.request("GET", url, headers=headers)

    print(response.text)

def task():
    url = "http://ccis-oamgate.cisdi.com.cn:7777/api/v1/todolist"

    payload = "{\"pageNum\":1,\"pageSize\":13,\"input\":\"\",\"type\":\"\",\"typeName\":\"\"}"
    headers = {
    'Accept': 'application/json, text/plain, */*',
    'Content-Type': 'application/Json; charset=UTF-8',
    'token': token,
    'Cookie': cookie
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    resp = json.loads(response.text)
    for item in resp['data']['csList']:
        getPage(item[''])


# loadExcel()

@app.route('/getData', methods=['GET'])
def getData():
    # 加载 Excel 文件
    workbook = load_workbook('main.xlsx')

    # 选择活动工作表
    sheet = workbook.active
    outputData = {
        "节点统计到期图纸": np.zeros(13, dtype=int).tolist(),
        "节点统计到期预算": np.zeros(13, dtype=int).tolist(),
        "节点统计到期采购": np.zeros(13, dtype=int).tolist(),
        "节点统计到期制造": np.zeros(13, dtype=int).tolist(),
        "节点统计到期检验": np.zeros(13, dtype=int).tolist(),
        "节点统计到期发运": np.zeros(13, dtype=int).tolist(),
        "节点统计超期图纸": np.zeros(13, dtype=int).tolist(),
        "节点统计超期预算": np.zeros(13, dtype=int).tolist(),
        "节点统计超期采购": np.zeros(13, dtype=int).tolist(),
        "节点统计超期制造": np.zeros(13, dtype=int).tolist(),
        "节点统计超期检验": np.zeros(13, dtype=int).tolist(),
        "节点统计超期发运": np.zeros(13, dtype=int).tolist(),
        "总已收账款": 0,
        "总应收账款": 0,
        "待收款性质": [0,0,0,0,0,0,0],
        "待收款金额": [0,0,0,0,0,0,0],
        "完成情况": {
            "进行中": 0,
            "发运完成": 0,
            "图纸下达": 0,
            "预算下达": 0,
            "成品检验完成": 0,
            "制造完成": 0,
            "采购合同完成": 0,
        },
        "项目状态": np.zeros(8, dtype=int).tolist()
    }
    def getDateStr (var):
        if (not var or var == '-'):
            # print('有不正确日期!')
            return 12
        if (not isinstance(var, datetime.datetime)):
            var = datetime.datetime.strptime(var, "%Y/%m/%d")
            
        if (var.year == 2024):
            return var.month - 1
        else:
            return 12
    def checkDate(date1, date2):
        if (date2 == '已完成' or date2 == '☑' or date1 == '-' or date2 == '-'):
            return False
        if (not date1):
            return False
        if (not date2):
            return False
        if (not isinstance(date1, datetime.datetime)):
            date1 = datetime.datetime.strptime(date1, "%Y/%m/%d")
        if (not isinstance(date2, datetime.datetime)):
            date2 = datetime.datetime.strptime(date2, "%Y/%m/%d")
        return [date1 < date2]
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if (row[0] != None):
            # 项目数量
            outputData["节点统计到期图纸"][getDateStr(row[16])] += 1
            outputData["节点统计到期预算"][getDateStr(row[19])] += 1
            outputData["节点统计到期采购"][getDateStr(row[22])] += 1
            outputData["节点统计到期制造"][getDateStr(row[25])] += 1
            outputData["节点统计到期检验"][getDateStr(row[28])] += 1
            outputData["节点统计到期发运"][getDateStr(row[31])] += 1
            # 延期数量
            if (checkDate(row[16], row[17])):
                outputData["节点统计超期图纸"][getDateStr(row[16])] += 1
            if (checkDate(row[19], row[20])):
                outputData["节点统计超期预算"][getDateStr(row[19])] += 1
            if (checkDate(row[22], row[23])):
                outputData["节点统计超期采购"][getDateStr(row[22])] += 1
            if (checkDate(row[25], row[26])):
                outputData["节点统计超期制造"][getDateStr(row[25])] += 1
            if (checkDate(row[28], row[29])):
                outputData["节点统计超期检验"][getDateStr(row[28])] += 1
            if (checkDate(row[31], row[32])):
                outputData["节点统计超期发运"][getDateStr(row[31])] += 1
            # 收款详情
            if (row[45]):
                outputData["总已收账款"] += int(row[45])
            if (row[39]):
                outputData["总应收账款"] += int(row[39])
            if (row[44]):
                if (row[44] == '预付款'):
                    outputData["待收款性质"][0] += 1
                    outputData["待收款金额"][0] += int(row[39])
                if (row[44] == '进度款'):
                    outputData["待收款性质"][1] += 1
                    outputData["待收款金额"][1] += int(row[39])
                if (row[44] == '提货款'):
                    outputData["待收款性质"][2] += 1
                    outputData["待收款金额"][2] += int(row[39])
                if (row[44] == '到货款'):
                    outputData["待收款性质"][3] += 1
                    outputData["待收款金额"][3] += int(row[39])
                if (row[44] == '调试款'):
                    outputData["待收款性质"][4] += 1
                    outputData["待收款金额"][4] += int(row[39])
                if (row[44] == '验收款'):
                    outputData["待收款性质"][5] += 1
                    outputData["待收款金额"][5] += int(row[39])
                if (row[44] == '质保金'):
                    outputData["待收款性质"][6] += 1
                    outputData["待收款金额"][6] += int(row[39])
            if (row[15]):
                outputData['项目状态'][0] += 1
                outputData["完成情况"][row[15]] += 1
            else:
                outputData['项目状态'][1] += 1
                outputData["完成情况"]["进行中"] += 1
            # 是否超期
            if (row[38]):
                outputData['项目状态'][2] += 1
            else:
                outputData['项目状态'][3] += 1
            # 罚款关注
            if (row[39]):
                outputData['项目状态'][4] += 1
            if (row[34]):
                outputData['项目状态'][5] += 1
    print(outputData)
    return json.dumps(outputData)

if __name__ == '__main__':
    app.run(debug=True)