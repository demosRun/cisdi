import re
import os
import time
import json
import base64
import datetime
import requests
from openpyxl import Workbook
import numpy as np
from openpyxl import load_workbook
from flask import Flask, send_file, request


# 定义要创建的文件夹路径
folder_path = './files'
 
# 如果文件夹不存在，则创建文件夹
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

def getFileList():
    print('开始尝试下载文件')
    url = "http://10.108.0.105:9999/asset-management-platform/api/apiDevelop/external/custom/api/oa/pa/zg/yjjh/list?appKey=2090000000008399559&appSecret=2206a222c4fd459f806a9255ee9332fa&enablePage=true&currPage=1&pageSize=10"
    payload = json.dumps({
    "project_number": 1
    })
    headers = {
    'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    return json.loads(response.text)

print('读取现有文档数据')
outPutData = []
workbook = load_workbook('file.xlsx')
sheet = workbook.active
for row in sheet.iter_rows(values_only=True):
    outPutData.append(list(row))

def loadExcel(file):
    print('加载文件: ' + file)
    # 加载 Excel 文件
    workbook = load_workbook(file)
    sheet = workbook.active
    项目号 = sheet.cell(row=3, column=4).value
    项目名称 = sheet.cell(row=2, column=4).value
    业主名称 = sheet.cell(row=2, column=11).value
    print(str(项目号) + ':' + 项目名称)

    # 获取从第四行开始的所有内容
    data = []
    # 定义替换函数
    def remove_zero_time(time):
        if (isinstance(time, datetime.datetime)):
            return time.strftime('%Y/%m/%d')
        return time
    产品名称 = ''
    总数量 = 0
    lastRow = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if (row[0]):
            lastRow = row
            if (产品名称 != ''):
                产品名称 += '/'
            产品名称 += row[7]
            if isinstance(row[8], str) and row[8].isdigit():
                总数量 += int(row[8])
            if isinstance(row[8], int):
                总数量 += int(row[8])
    lestTemp = [项目名称, re.findall(r'\d+', 项目名称)[-1], 项目号, "", re.findall(r'\d+', 项目名称)[0], 业主名称, 产品名称, 总数量, None, None, None, lastRow[12], None,None,None,lastRow[15],None,None,lastRow[16],None,None, lastRow[17],None,None,lastRow[18],None,None,lastRow[19],None,None, lastRow[20]]
    lestTemp = list(map(remove_zero_time, lestTemp))
    data.append(lestTemp)
    return data


def outPutFile(requestid):
    url = "http://oa.cisdi.com.cn:8080/api/sdzb/getFileDate"
    payload = json.dumps({
        "requestid": requestid
    })
    headers = {
    'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    dataTemp = json.loads(response.text)
    # 假设这是你的Base64编码的字符串
    base64_string = dataTemp['resultData'][0]['base64']
    
    
    
    # 指定要创建的文件名
    file_name = dataTemp['resultData'][0]['filename']
    if ('项目立项审批流程' in file_name):
        # 将解码后的数据写入文件
        if (not os.path.exists('./files/' + file_name)):
            # 解码Base64字符串
            decoded_data = base64.b64decode(base64_string)
            with open('./files/' + file_name, 'wb') as file:
                file.write(decoded_data)
                time.sleep(1)
                outPutData.extend(loadExcel('./files/' + file_name))

# 网络请求
fileList = getFileList()
print(fileList['data'])
fileList = fileList['data']['rowList']
for item in fileList:
    print(item['requestname'])
    outPutFile(item['requestid'])

# 保存为CSV文件
print(f"保存到xlsx文件")
# 创建一个新的Workbook对象
wb = Workbook()
ws = wb.active  # 获取活动的工作表

# 将数组中的每一行写入工作表
for row in outPutData:
    ws.append(row)

# 保存为.xlsx文件
wb.save('file.xlsx')

